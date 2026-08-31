import json as _json
_M=_json.load(open("meta.json"))
DATE_CUR=_M["DATE_CUR"]; DATE_PREV=_M["DATE_PREV"]; LBL_CUR=_M["LBL_CUR"]; LBL_PREV=_M["LBL_PREV"]; OUT=_M["OUT"]
FII_HDR=_M.get("FII_HDR","FII Δ Qtr %"); DII_HDR=_M.get("DII_HDR","DII Δ 3Y %")
MON_PREV=LBL_PREV.split("-")[-1]; MON_CUR=LBL_CUR.split("-")[-1]
import os as _os; HAS_WL=_os.path.exists("wl2.pkl")
import pandas as pd, numpy as np, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.styles.colors import Color
from openpyxl.utils import get_column_letter

M=pd.read_pickle("master.pkl"); meta=json.load(open("meta.json"))
book=meta["book"]; FULL_JUN=meta["FULL_JUN"]; EXITED_JUN=meta["EXITED_JUN"]; HELD_JUN=meta["jun_held"]
AR="Arial"; NAVY="1F3864"; BLUE="2E5496"; green="006100"; red="9C0006"
thin=Side(style="thin",color="BFBFBF"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
MAF={">50 & >200 DMA":Color(theme=6,tint=0.8),"mixed DMA":Color(theme=9,tint=0.8),"<both DMA":Color(theme=5,tint=0.8)}
VC={"GREEN ZONE":"C6EFCE","MOMENTUM PLAY":"DDEBF7","VALUE OPPORTUNITY":"FFF2CC","AVOID":"E7E6E6","EXIT":"FCE4E4","EXIT (Stage 1)":"F8CBCB"}
RF={"ADD":"C6EFCE","BUY":"C6EFCE","BUY (momentum)":"C6EFCE","HOLD/ADD":"D9EAD3","HOLD (ride)":"DDEBF7","HOLD":"FFFFFF","HOLD (value)":"FFF2CC","ACCUMULATE (weak mom)":"FFF2CC","WATCH":"F2F2F2","WATCH (value, weak mom)":"F2F2F2","TRIM (fading)":"FCE4D6","TRIM":"FCE4D6","TRIM/EXIT":"F8CBCB","EXIT":"F8CBCB","AVOID":"E7E6E6","REVIEW":"FFFFFF"}
RC={"ADD":"006100","BUY":"006100","BUY (momentum)":"006100","HOLD/ADD":"006100","HOLD (ride)":"1F4E79","TRIM (fading)":"9C5700","TRIM":"9C5700","TRIM/EXIT":"9C0006","EXIT":"9C0006","AVOID":"808080"}
def AT(a):
    a=str(a)
    for k,v in {"K-ICICI":"KI","S-ICICI":"SI","K-Kite":"KZ"}.items(): a=a.replace(k,v)
    return a
def N(v): return "" if (v is None or (isinstance(v,float) and pd.isna(v))) else v
wb=Workbook()

# ================= SHEET 1: Portfolio + Factors =================
ws=wb.active; ws.title="Portfolio + Factors"; ws.sheet_view.showGridLines=False
ws["A1"]=f"PORTFOLIO + FACTORS  —  {DATE_CUR}"; ws["A1"].font=Font(name=AR,bold=True,size=13,color=NAVY)
ws["A2"]=(f"{DATE_PREV} → {DATE_CUR}. Per-row {LBL_PREV} value covers only positions still held; ₹{EXITED_JUN/1e5:.1f}L of holdings have since exited "
          f"& new buys have no prior-week value. FULL BOOK: {LBL_PREV} ₹{FULL_JUN/1e5:.1f}L → {LBL_CUR} ₹{book/1e5:.1f}L = ₹{(book-FULL_JUN)/1e5:+.1f}L.")
ws["A2"].font=Font(name=AR,italic=True,size=8.5,color="808080")
bands=[("POSITION",1,10,BLUE),("FUNDAMENTALS",11,20,"548235"),("TECHNICALS",21,25,NAVY)]
for lbl,c1,c2,col in bands:
    ws.merge_cells(start_row=4,start_column=c1,end_row=4,end_column=c2)
    for cc in range(c1,c2+1):
        x=ws.cell(4,cc); x.fill=PatternFill("solid",start_color=col); x.border=bd
    x=ws.cell(4,c1,lbl); x.font=Font(name=AR,bold=True,size=9,color="FFFFFF"); x.alignment=Alignment(horizontal="center")
pcols=["Company",f"{MON_PREV} Qty",f"{MON_CUR} Qty","Δ Qty",f"{LBL_PREV} Value",f"{LBL_CUR} Value","Δ Value (₹)","% Chg","P&L %","% of Port",
       "Qtr Sales %","Qtr Profit %","Sales 3Y %","Sales 5Y %","OPM %","PE","ROCE %","Pledge %",FII_HDR,DII_HDR,
       "RSI","% from 52wH","Wk Ret %","1M Ret %","MA Position"]
W=[26,8,8,7,13,13,13,8,8,8,10,10,9,9,8,8,9,9,9,9,7,10,9,9,14]
for i,h in enumerate(pcols,1):
    c=ws.cell(5,i,h); c.font=Font(name=AR,bold=True,color="FFFFFF",size=8); c.fill=PatternFill("solid",start_color=NAVY)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=bd
    ws.column_dimensions[get_column_letter(i)].width=W[i-1]
SIGN={"% Chg","P&L %","Qtr Sales %","Qtr Profit %","Sales 3Y %","Sales 5Y %","1M Ret %","Wk Ret %","% from 52wH",FII_HDR,DII_HDR}
CURC={f"{LBL_PREV} Value",f"{LBL_CUR} Value","Δ Value (₹)"}
def putP(r,vals,fill=None,bold=False,marow=None):
    for ci,v in enumerate(vals,1):
        c=ws.cell(r,ci,N(v)); c.border=bd; c.font=Font(name=AR,size=8.5,bold=bold)
        cn=pcols[ci-1]
        if marow in MAF: c.fill=PatternFill(patternType="solid",fgColor=MAF[marow])
        elif fill: c.fill=PatternFill("solid",start_color=fill)
        if cn in CURC and isinstance(v,(int,float)) and pd.notna(v): c.number_format="#,##0;[Red]-#,##0"
        if cn in (f"{MON_PREV} Qty",f"{MON_CUR} Qty","Δ Qty") and isinstance(v,(int,float)) and pd.notna(v): c.number_format="#,##0"; c.alignment=Alignment(horizontal="right")
        if cn=="PE" and isinstance(v,(int,float)) and pd.notna(v): c.number_format="0.0"
        if cn in SIGN and isinstance(v,(int,float)) and pd.notna(v):
            c.number_format="+0.0;[Red]-0.0"
            if cn in ("% Chg","P&L %",FII_HDR,DII_HDR): c.font=Font(name=AR,size=8.5,bold=bold,color=green if v>=0 else red)
        elif cn in ("% of Port","OPM %","ROCE %","Pledge %","RSI") and isinstance(v,(int,float)) and pd.notna(v): c.number_format="0.0"
        if cn in ("RSI","PE","% of Port","MA Position"): c.alignment=Alignment(horizontal="center")
GRP={"Gold/Commodities":("GOLD / COMMODITIES","806000"),"ETF (Index)":("ETFs (INDEX)","1F4E79"),"Equity":("EQUITY","548235")}
r=6
for grp in ["Gold/Commodities","ETF (Index)","Equity"]:
    sub=M[M.Class==grp].sort_values("JulVal",ascending=False); title,hc=GRP[grp]
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=23)
    x=ws.cell(r,1,title); x.font=Font(name=AR,bold=True,size=9,color="FFFFFF"); x.fill=PatternFill("solid",start_color=hc); x.alignment=Alignment(indent=1); r+=1
    for _,m in sub.iterrows():
        iseq=(grp=="Equity")
        vals=[m["Company"],N(m["JunQty"]),m["JulQty"],N(m["Dqty"]),N(m["JunVal"]) if m["JunVal"]>0 else None,m["JulVal"],
              N(m["Dval"]),N(m["PctChg"]),N(m["PnLpct"]),m["PortPct"],
              N(m["QSales"]) if iseq else None,N(m["QProfit"]) if iseq else None,N(m["S3"]) if iseq else None,N(m["S5"]) if iseq else None,
              N(m["OPM"]) if iseq else None,N(m["PE"]) if iseq else None,N(m["ROCE"]) if iseq else None,N(m["Pledge"]) if iseq else None,
              N(m["FIIchg"]) if iseq else None,N(m["DIIchg"]) if iseq else None,
              N(m["RSI"]) if iseq else None,N(m["Dfh"]) if iseq else None,N(m["Wk"]) if iseq else None,N(m["M1"]) if iseq else None,m["MA"] if iseq else ""]
        putP(r,vals,marow=(m["MA"] if iseq else None)); r+=1
    # subtotal
    _pl=np.average(sub["PnLpct"].fillna(0),weights=sub["JulVal"]) if sub["JulVal"].sum()>0 else 0
    _jn=sub["JunVal"].fillna(0).sum(); _jl=sub["JulVal"].sum(); _cg=((_jl-_jn)/_jn*100) if _jn>0 else None
    st=["  Sub-total — "+title.title(),"","","",(_jn if _jn>0 else None),_jl,_jl-_jn,_cg,_pl,_jl/book*100]+[None]*15
    putP(r,st,fill="F2F2F2",bold=True); r+=1
# reconciliation
r+=1
def recon(r,label,jun,jul,fill,bold=True):
    ws.merge_cells(start_row=r,start_column=1,end_row=r,end_column=4)
    x=ws.cell(r,1,label); x.font=Font(name=AR,bold=bold,color="FFFFFF" if fill==NAVY else "000000"); x.fill=PatternFill("solid",start_color=fill); x.alignment=Alignment(indent=1)
    for cc in range(1,24): ws.cell(r,cc).fill=PatternFill("solid",start_color=fill); ws.cell(r,cc).border=bd
    ws.cell(r,1,label).font=Font(name=AR,bold=bold,color="FFFFFF" if fill==NAVY else "000000")
    if jun is not None: c=ws.cell(r,5,jun); c.number_format="#,##0"; c.font=Font(name=AR,bold=bold,color="FFFFFF" if fill==NAVY else "000000")
    if jul is not None: c=ws.cell(r,6,jul); c.number_format="#,##0"; c.font=Font(name=AR,bold=bold,color="FFFFFF" if fill==NAVY else "000000")
    if jun is not None and jul is not None:
        c=ws.cell(r,7,jul-jun); c.number_format="#,##0;[Red]-#,##0"; c.font=Font(name=AR,bold=bold,color="FFFFFF" if fill==NAVY else "000000")
recon(r,"GRAND TOTAL (held positions)",HELD_JUN,book,"D9E1F2"); r+=1
recon(r,f"+ Exited since {LBL_PREV} ({MON_PREV} value)",EXITED_JUN,None,"FCE4E4"); r+=1
recon(r,"= FULL BOOK",FULL_JUN,book,NAVY); r+=1
ws.freeze_panes="B6"

print("Sheet1 Portfolio+Factors done, last row", r)

# ================= SHEET 2: Equity Master =================
we=wb.create_sheet("Equity Master"); we.sheet_view.showGridLines=False
we["A1"]=f"EQUITY MASTER  —  {DATE_CUR}  (rows colour-coded by MA Position)"; we["A1"].font=Font(name=AR,bold=True,size=13,color=NAVY)
eq=M[M.Class=="Equity"].sort_values("JulVal",ascending=False).reset_index(drop=True)
we["A2"]=f"{len(eq)} held equities · {(eq.Verdict=='GREEN ZONE').sum()} GREEN ZONE · {eq.Reco.isin(['ADD','BUY','HOLD/ADD']).sum()} ADD/BUY. 7 unscored (not in Screener) shown with blank CONVICTION-6."
we["A2"].font=Font(name=AR,italic=True,size=8.5,color="808080")
ebands=[("POSITION",1,16,BLUE),("TECHNICALS",17,20,NAVY),("FUNDAMENTALS",21,30,"548235"),("CONVICTION-6",31,40,"7030A0")]
for lbl,c1,c2,col in ebands:
    we.merge_cells(start_row=4,start_column=c1,end_row=4,end_column=c2)
    for cc in range(c1,c2+1):
        x=we.cell(4,cc); x.fill=PatternFill("solid",start_color=col); x.border=bd
    x=we.cell(4,c1,lbl); x.font=Font(name=AR,bold=True,size=9,color="FFFFFF"); x.alignment=Alignment(horizontal="center")
ecols=["Company","Account(s)","NSE","Recommendation","TV","Cup & Handle",f"{MON_PREV} Qty",f"{MON_CUR} Qty","Δ Qty",f"{LBL_PREV} Val",f"{LBL_CUR} Val","Δ Value (₹)","% Chg","P&L %","% from 52wH","% of Port",
       "RSI","Wk Ret %","1M Ret %","MA Position",
       "Qtr Sales %","Qtr Profit %","Sales 3Y","Sales 5Y","OPM %","PE","ROCE %","Pledge %",FII_HDR,DII_HDR,
       "Fund/40","FG","Tech/8","TG","Grand/48","Verdict","Conv/100","Tier","Wt","PEG"]
WE=[24,9,11,14,9,14,7,7,7,12,12,12,8,8,10,8,7,8,8,14,9,9,8,8,8,7,8,8,9,9,8,5,7,8,8,14,8,7,7,7]
for i,h in enumerate(ecols,1):
    c=we.cell(5,i,h); c.font=Font(name=AR,bold=True,color="FFFFFF",size=8); c.fill=PatternFill("solid",start_color=NAVY)
    c.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); c.border=bd
    we.column_dimensions[get_column_letter(i)].width=WE[i-1]
ESIGN={"% Chg","P&L %","Qtr Sales %","Qtr Profit %","Sales 3Y","Sales 5Y","1M Ret %","Wk Ret %","% from 52wH",FII_HDR,DII_HDR}
ECUR={f"{LBL_PREV} Val",f"{LBL_CUR} Val","Δ Value (₹)"}
import re as _re_em, os as _os_em
def _emkey(nse,comp=""):
    b=str(nse).strip().upper()
    if b in ("","NAN","NONE"): b=str(comp).strip().upper()
    b=_re_em.sub(r"-(BE|BZ|BL|SM|P\d+|N\d+)$","",b); return b.replace(" ","").replace(".","").replace("&","")
GROUP_OF={}; CUP_TIER={}
if _os_em.path.exists("wl2.pkl"):
    _wlx=pd.read_pickle("wl2.pkl"); GROUP_OF={k:str(g).title() for k,g in zip(_wlx["key"],_wlx["Group"])}
if _os_em.path.exists("cup.pkl"):
    _cupx=pd.read_pickle("cup.pkl")
    if len(_cupx): CUP_TIER={_emkey(rw["NSE"],rw["Company"]):rw["Tier"] for _,rw in _cupx.iterrows()}
r=6
for _,m in eq.iterrows():
    reco=str(m["Reco"]) if isinstance(m["Reco"],str) else ""
    tvg=GROUP_OF.get(m["key"],""); cupt=CUP_TIER.get(m["key"],"")
    vals=[m["Company"],AT(m["Acct"]),m["NSE"],reco,tvg,cupt,N(m["JunQty"]),m["JulQty"],N(m["Dqty"]),
          N(m["JunVal"]) if m["JunVal"]>0 else None,m["JulVal"],N(m["Dval"]),N(m["PctChg"]),N(m["PnLpct"]),N(m["Dfh"]),m["PortPct"],
          N(m["RSI"]),N(m["Wk"]),N(m["M1"]),m["MA"],
          N(m["QSales"]),N(m["QProfit"]),N(m["S3"]),N(m["S5"]),N(m["OPM"]),N(m["PE"]),N(m["ROCE"]),N(m["Pledge"]),N(m["FIIchg"]),N(m["DIIchg"]),
          N(m["Fund"]),N(m["FG"]),N(m["Tech"]),N(m["TG"]),N(m["Grand"]),N(m["Verdict"]),N(m["Conv"]),N(m["Tier"]),N(m["Wt"]),N(m["PEG"])]
    ma=m["MA"] if m["MA"] in MAF else None
    for ci,v in enumerate(vals,1):
        c=we.cell(r,ci,N(v)); c.border=bd; c.font=Font(name=AR,size=8.5); cn=ecols[ci-1]
        if ma: c.fill=PatternFill(patternType="solid",fgColor=MAF[ma])
        if cn in ECUR and isinstance(v,(int,float)) and pd.notna(v): c.number_format="#,##0;[Red]-#,##0"
        if cn in (f"{MON_PREV} Qty",f"{MON_CUR} Qty","Δ Qty") and isinstance(v,(int,float)) and pd.notna(v): c.number_format="#,##0"; c.alignment=Alignment(horizontal="right")
        if cn in ESIGN and isinstance(v,(int,float)) and pd.notna(v):
            c.number_format="+0.0;[Red]-0.0"
            if cn in ("% Chg","P&L %",FII_HDR,DII_HDR): c.font=Font(name=AR,size=8.5,color=green if v>=0 else red)
        elif cn in ("% of Port","OPM %","ROCE %","Pledge %","RSI") and isinstance(v,(int,float)) and pd.notna(v): c.number_format="0.0"
        if cn in ("PE","PEG","Fund/40","Tech/8","Grand/48","Conv/100") and isinstance(v,(int,float)) and pd.notna(v): c.number_format="0.0" if cn not in("PEG",) else "0.00"
        if cn=="Recommendation" and reco: c.font=Font(name=AR,size=8.5,bold=True,color=RC.get(reco,"808080"))
        if cn=="TV" and v:
            c.alignment=Alignment(horizontal="center"); c.font=Font(name=AR,size=8,bold=True,color={"Blue":"1F4E79","Orange":"C55A11","Green":"548235"}.get(v,"808080"))
        if cn=="Cup & Handle" and v:
            c.alignment=Alignment(horizontal="center"); c.font=Font(name=AR,size=8,bold=True,color="1F4E79" if v=="Breakout-ready" else "9C5700")
            c.fill=PatternFill("solid",start_color="C6EFCE" if v=="Breakout-ready" else "FFF2CC")
        if cn in ("NSE","FG","TG","Tier","MA Position","% of Port","Verdict","Conv/100","Fund/40","Tech/8","Grand/48","PEG","Wt"): c.alignment=Alignment(horizontal="center")
    r+=1
we.freeze_panes="G6"


# ================= SHEET 3: Industry (flat, filterable; change subtotals) =================
wi=wb.create_sheet("Industry"); wi.sheet_view.showGridLines=False
wi["A1"]=f"INDUSTRY BREAKDOWN  —  held equities, {DATE_CUR}"; wi["A1"].font=Font(name=AR,bold=True,size=13,color=NAVY)
wi["A2"]="Flat table — use the Industry filter (row 5). Each industry has a sub-total incl. Δ Val & % Chg vs prior week. Technicals + CONVICTION-6 per company."
wi["A2"].font=Font(name=AR,italic=True,size=8.5,color="808080")
icols=["Industry","Company","NSE","Acct",f"{LBL_CUR} Val","Δ Val","% Chg","% Book","P&L %","Wk %","1M %","RSI","MA Position","Tech/8","TG","Grand/48","Verdict","Conv/100","Reco"]
WI=[22,22,10,7,12,11,8,7,8,7,7,6,14,7,6,8,15,8,13]
ibands=[("POSITION",1,11,BLUE),("TECHNICALS",12,13,NAVY),("CONVICTION-6",14,19,"7030A0")]
for lbl,c1,c2,col in ibands:
    wi.merge_cells(start_row=4,start_column=c1,end_row=4,end_column=c2)
    for cc in range(c1,c2+1):
        x=wi.cell(4,cc); x.fill=PatternFill("solid",start_color=col); x.border=bd
    wi.cell(4,c1,lbl).font=Font(name=AR,bold=True,size=9,color="FFFFFF"); wi.cell(4,c1).alignment=Alignment(horizontal="center")
for i,h in enumerate(icols,1):
    c=wi.cell(5,i,h); c.font=Font(name=AR,bold=True,color="FFFFFF",size=8); c.fill=PatternFill("solid",start_color=NAVY)
    c.alignment=Alignment(horizontal="center" if i>=5 else "left",vertical="center",wrap_text=True); c.border=bd
    wi.column_dimensions[get_column_letter(i)].width=WI[i-1]
ISIGN={"Δ Val","% Chg","Wk %","P&L %","1M %"}
def iput(r,vals,ma=None,fill=None,bold=False):
    for ci,v in enumerate(vals,1):
        c=wi.cell(r,ci,N(v)); c.border=bd; c.font=Font(name=AR,size=8.5,bold=bold); cn=icols[ci-1]
        if fill: c.fill=PatternFill("solid",start_color=fill)
        elif ma and cn!="Industry": c.fill=PatternFill(patternType="solid",fgColor=MAF[ma])
        if cn in (f"{LBL_CUR} Val","Δ Val"): c.number_format="#,##0;[Red]-#,##0"; c.alignment=Alignment(horizontal="right")
        if cn=="% Book": c.number_format="0.0"; c.alignment=Alignment(horizontal="center")
        if cn in ("% Chg","Wk %","P&L %","1M %"):
            c.number_format="+0.0;[Red]-0.0"; c.alignment=Alignment(horizontal="right")
            if isinstance(v,(int,float)) and pd.notna(v): c.font=Font(name=AR,size=8.5,bold=bold,color=green if v>=0 else red)
        if cn in ("RSI","Tech/8","Grand/48","Conv/100","TG","MA Position"): c.alignment=Alignment(horizontal="center")
        if cn in ("RSI","Grand/48") and isinstance(v,(int,float)) and pd.notna(v): c.number_format="0.0"
        if cn=="Verdict":
            fl=VC.get(str(v))
            if fl and not fill: c.fill=PatternFill("solid",start_color=fl)
            c.alignment=Alignment(horizontal="center"); c.font=Font(name=AR,size=8,bold=bold)
        if cn=="Reco" and isinstance(v,str) and v: c.font=Font(name=AR,size=8,bold=True,color=RC.get(v,"808080"))
        if cn=="NSE": c.font=Font(name=AR,size=8,color="595959")
        if cn=="Industry": c.font=Font(name=AR,size=8.5,bold=bold,color="1F3864")
        if cn=="Company" and bold: c.font=Font(name=AR,size=8.5,bold=True)
eqi=eq.copy(); eqi["Industry"]=eqi["Industry"].fillna("(unclassified)")
indtot=eqi.groupby("Industry")["JulVal"].sum().sort_values(ascending=False)
r=6
for ind in indtot.index:
    sdf=eqi[eqi.Industry==ind].sort_values("JulVal",ascending=False)
    for _,m in sdf.iterrows():
        ma=m["MA"] if m["MA"] in MAF else None
        iput(r,[ind,m["Company"],m["NSE"],AT(m["Acct"]),m["JulVal"],N(m["Dval"]),N(m["PctChg"]),m["PortPct"],N(m["PnLpct"]),N(m["Wk"]),N(m["M1"]),
                N(m["RSI"]),m["MA"],N(m["Tech"]),N(m["TG"]),N(m["Grand"]),N(m["Verdict"]),N(m["Conv"]),N(m["Reco"])],ma=ma)
        r+=1
    _jn=sdf["JunVal"].fillna(0).sum(); _jl=sdf["JulVal"].sum(); _dv=sdf["Dval"].fillna(0).sum()
    _cg=((_jl-_jn)/_jn*100) if _jn>0 else None
    _pl=np.average(sdf.PnLpct.fillna(0),weights=sdf.JulVal) if _jl>0 else 0
    _wk=np.average(sdf.Wk.fillna(0),weights=sdf.JulVal) if _jl>0 else 0
    iput(r,[ind,f"Sub-total — {ind}","","",_jl,_dv,_cg,_jl/book*100,_pl,_wk,"","","","","","","","",""],fill="EDEDED",bold=True)
    r+=1
# grand total
_jl=eqi.JulVal.sum(); _dv=eqi["Dval"].fillna(0).sum(); _jn=eqi["JunVal"].fillna(0).sum()
_cg=((_jl-_jn)/_jn*100) if _jn>0 else None; _pl=np.average(eqi.PnLpct.fillna(0),weights=eqi.JulVal); _wk=np.average(eqi.Wk.fillna(0),weights=eqi.JulVal)
for cc in range(1,21): wi.cell(r,cc).fill=PatternFill("solid",start_color="1F4E79"); wi.cell(r,cc).border=bd
tot=["","TOTAL EQUITY","","",_jl,_dv,_cg,_jl/book*100,_pl,_wk,"","","","","","","","",""]
for ci,v in enumerate(tot,1):
    c=wi.cell(r,ci,N(v)); c.font=Font(name=AR,bold=True,color="FFFFFF",size=8.5)
    if icols[ci-1] in (f"{LBL_CUR} Val","Δ Val"): c.number_format="#,##0;[Red]-#,##0"; c.alignment=Alignment(horizontal="right")
    if icols[ci-1] in ("% Chg","P&L %"): c.number_format="+0.0;[Red]-0.0"; c.alignment=Alignment(horizontal="right")
    if icols[ci-1]=="% Book": c.number_format="0.0"; c.alignment=Alignment(horizontal="center")
lastrow=r
wi.auto_filter.ref=f"A5:{get_column_letter(len(icols))}{lastrow}"
wi.freeze_panes="A6"

if HAS_WL:
    # ================= SHEET 6: TradingView Watchlist =================
    wv=wb.create_sheet("TradingView Watchlist"); wv.sheet_view.showGridLines=False
    import re as _re
    wl=pd.read_pickle("wl2.pkl"); scr2=pd.read_csv("screener.csv"); scr2.columns=[c.strip() for c in scr2.columns]
    def _K(s):
        b=str(s).strip().upper(); b=_re.sub(r"-(BE|BZ|BL|SM|P\d+|N\d+)$","",b); return b.replace(" ","").replace(".","").replace("&","")
    scr2["key"]=scr2["NSE Code"].map(_K); s2=scr2.set_index("key")
    scd=pd.read_pickle("scored.pkl"); scd["key"]=scd.apply(lambda r:_K(r.get("NSE") if pd.notna(r.get("NSE")) else r["Name"]),axis=1); sc2=scd.set_index("key")
    def sv(k,idx,c):
        try:
            v=idx.loc[k,c]; return v if pd.notna(v) else ""
        except: return ""
    wv["A1"]=f"TRADINGVIEW WATCHLIST — Blue / Orange / Green  ({DATE_CUR})  — rows colour-coded by MA Position"
    wv["A1"].font=Font(name=AR,bold=True,size=13,color=NAVY)
    wv["A2"]=(f"{len(wl)} names · {int(wl['Held'].sum())} held · {int(wl['Scored'].sum())} scored. "
              f"Held? and CONVICTION-6 recos shown where the name is in the held-Screener export; "
              f"{int((~wl['Scored']).sum())} watch-only names need a watchlist Screener CSV to score.")
    wv["A2"].font=Font(name=AR,italic=True,size=8.5,color="808080")
    vcols=["Group","Symbol","Company","Held?","Last","RSI","Wk %","1M %","MA Position","Fund/40","FG","Tech/8","TG","Verdict","Conv","Mom","Recommendation"]
    VW=[9,12,24,7,9,6,7,7,14,7,5,7,7,15,6,6,16]
    vbands=[("LIST",1,5,BLUE),("TECHNICALS",6,9,NAVY),("CONVICTION-6",10,17,"7030A0")]
    for lbl,c1,c2,col in vbands:
        wv.merge_cells(start_row=4,start_column=c1,end_row=4,end_column=c2)
        for cc in range(c1,c2+1):
            x=wv.cell(4,cc); x.fill=PatternFill("solid",start_color=col); x.border=bd
        wv.cell(4,c1,lbl).font=Font(name=AR,bold=True,size=9,color="FFFFFF"); wv.cell(4,c1).alignment=Alignment(horizontal="center")
    for i,h in enumerate(vcols,1):
        c=wv.cell(5,i,h); c.font=Font(name=AR,bold=True,color="FFFFFF",size=8); c.fill=PatternFill("solid",start_color=NAVY)
        c.alignment=Alignment(horizontal="center" if i>=4 else "left",vertical="center",wrap_text=True); c.border=bd
        wv.column_dimensions[get_column_letter(i)].width=VW[i-1]
    GT={"BLUE":"1F4E79","ORANGE":"C55A11","GREEN":"548235"}
    GLABEL={"BLUE":"BLUE — Weekly Bases","ORANGE":"ORANGE — KR","GREEN":"GREEN — To Study"}
    wl=wl.copy(); wl["ord"]=wl["Group"].map({"BLUE":0,"ORANGE":1,"GREEN":2})
    # within group: scored first by Conv desc, then unscored alpha
    wl["ConvV"]=wl["key"].map(lambda k: (sc2.loc[k,"Conv"] if k in sc2.index and pd.notna(sc2.loc[k,"Conv"]) else -1))
    wl=wl.sort_values(["ord","Scored","ConvV"],ascending=[True,False,False])
    r=6
    for grp in ["BLUE","ORANGE","GREEN"]:
        sub=wl[wl.Group==grp]
        wv.merge_cells(start_row=r,start_column=1,end_row=r,end_column=17)
        h=wv.cell(r,1,f"{GLABEL[grp]}  —  {len(sub)} names · {int(sub['Held'].sum())} held")
        h.font=Font(name=AR,bold=True,size=9,color="FFFFFF"); h.fill=PatternFill("solid",start_color=GT[grp]); h.alignment=Alignment(indent=1); r+=1
        for _,x in sub.iterrows():
            k=x["key"]; scored=x["Scored"]
            ma=sv(k,sc2,"MA Position") if scored else ""
            maf=ma if ma in MAF else None
            reco=sv(k,sc2,"Reco") if scored else ""
            vals=[grp,x["Symbol"],x["Company"],"Yes" if x["Held"] else "No",x["Last"],
                  sv(k,s2,"RSI"),sv(k,s2,"Return over 1week"),sv(k,s2,"Return over 1month"),ma,
                  sv(k,sc2,"Fund"),sv(k,sc2,"FGrade"),sv(k,sc2,"Tech"),sv(k,sc2,"TGrade"),
                  sv(k,sc2,"Verdict"),sv(k,sc2,"Conv"),sv(k,sc2,"Mom"),reco]
            for ci,v in enumerate(vals,1):
                c=wv.cell(r,ci,N(v) if not isinstance(v,str) else v); c.border=bd; c.font=Font(name=AR,size=8.5); cn=vcols[ci-1]
                if maf and cn not in ("Group",): c.fill=PatternFill(patternType="solid",fgColor=MAF[maf])
                if cn=="Group":
                    c.fill=PatternFill("solid",start_color=GT[grp]); c.font=Font(name=AR,size=7.5,bold=True,color="FFFFFF"); c.alignment=Alignment(horizontal="center")
                if cn=="Symbol": c.font=Font(name=AR,size=8,bold=True,color="1F3864")
                if cn=="Held?":
                    c.alignment=Alignment(horizontal="center")
                    c.font=Font(name=AR,size=8,bold=True,color=(green if x["Held"] else "808080"))
                if cn=="Last" and isinstance(v,(int,float)): c.number_format="#,##0.00"; c.alignment=Alignment(horizontal="right")
                if cn in ("RSI","Conv","Mom","Fund/40","Tech/8") and isinstance(v,(int,float)): c.number_format="0.0" if cn in("RSI","Conv","Mom","Fund/40") else "0"; c.alignment=Alignment(horizontal="center")
                if cn in ("Wk %","1M %") and isinstance(v,(int,float)):
                    c.number_format="+0.0;[Red]-0.0"; c.alignment=Alignment(horizontal="right"); c.font=Font(name=AR,size=8.5,color=green if v>=0 else red)
                if cn in ("FG","TG","MA Position"): c.alignment=Alignment(horizontal="center")
                if cn=="Verdict" and isinstance(v,str) and v:
                    fl=VC.get(v)
                    if fl: c.fill=PatternFill("solid",start_color=fl)
                    c.alignment=Alignment(horizontal="center"); c.font=Font(name=AR,size=7.5)
                if cn=="Mom" and isinstance(v,(int,float)):
                    mc=green if v>=70 else "9C5700" if v>=45 else red; c.font=Font(name=AR,size=8.5,color=mc)
                if cn=="Recommendation" and isinstance(v,str) and v: c.font=Font(name=AR,size=8,bold=True,color=RC.get(v,"808080"))
            r+=1
    r+=1; wv.merge_cells(start_row=r,start_column=1,end_row=r,end_column=17)
    unsc=", ".join(sorted(wl[~wl.Scored]["Symbol"].tolist())[:60])
    wv.cell(r,1,"Watch-only (not held, not in Screener) — provide a watchlist Screener CSV to score & recommend these.").font=Font(name=AR,italic=True,size=8,color="9C5700")
    wv.freeze_panes="A6"

def flatten_and_filter(wb):
    from copy import copy
    FROW={"Portfolio + Factors":5,"Equity Master":5,"Industry":5,"Cup & Handle":4,"Recommendations":7,"TradingView Watchlist":5}
    for ws in wb.worksheets:
        for rng in list(ws.merged_cells.ranges):
            tl=ws.cell(rng.min_row,rng.min_col); fill=copy(tl.fill)
            ws.unmerge_cells(str(rng))
            for rr in range(rng.min_row,rng.max_row+1):
                for cc in range(rng.min_col,rng.max_col+1):
                    c=ws.cell(rr,cc)
                    if not (rr==rng.min_row and cc==rng.min_col): c.fill=copy(fill)
        hr=FROW.get(ws.title)
        if hr and ws.max_row>=hr:
            ws.auto_filter.ref=f"A{hr}:{get_column_letter(ws.max_column)}{ws.max_row}"
flatten_and_filter(wb)
out=OUT
wb.save(out); print("SAVED:",out); print("Sheets:",wb.sheetnames)
