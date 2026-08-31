#!/usr/bin/env python3
import sys, re
from copy import copy
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
SRC=sys.argv[1] if len(sys.argv)>1 else "in.xlsx"; DST=sys.argv[2] if len(sys.argv)>2 else SRC
HDR_ROW=5
TOTAL_RE=re.compile(r"(sub-?total|grand total|total equity|^\s*total)",re.I); NOTE_RE=re.compile(r"(exited since|watch-only|not in screener)",re.I)
NAVY="1F3864"; thin=Side(style="thin",color="D9D9D9"); BORDER=Border(left=thin,right=thin,top=thin,bottom=thin)
def rc(ws,r): return [ws.cell(r,c) for c in range(1,ws.max_column+1)]
def filled(ws,r): return sum(1 for c in rc(ws,r) if c.value not in (None,""))
def classify(ws,r):
    if filled(ws,r)==0: return "blank"
    a=str(ws.cell(r,1).value or ""); b=str(ws.cell(r,2).value or "")
    if TOTAL_RE.search(a) or TOTAL_RE.search(b): return "total"
    if NOTE_RE.search(a): return "note"
    if filled(ws,r)<=2: return "banner"
    return "data"
def cpc(src,dst):
    dst.value=src.value
    if src.has_style:
        dst.font=copy(src.font);dst.fill=copy(src.fill);dst.border=copy(src.border);dst.alignment=copy(src.alignment);dst.number_format=src.number_format
def flatten(ws,bcol=None):
    maxc=ws.max_column
    head=[[ws.cell(r,c) for c in range(1,maxc+1)] for r in range(1,HDR_ROW+1)]
    data=[];tail=[];banner=None;seen=[]
    for r in range(HDR_ROW+1,ws.max_row+1):
        k=classify(ws,r)
        if k=="banner": banner=str(ws.cell(r,1).value).split("—")[0].strip(); seen.append(banner)
        elif k=="data": data.append((rc(ws,r),banner))
        elif k in ("total","note"): tail.append(rc(ws,r))
    ins=bcol is not None and seen; width=maxc+(1 if ins else 0)
    ws.delete_rows(1,ws.max_row)
    for ri,cells in enumerate(head,start=1):
        off=0
        if ins:
            t=ws.cell(ri,1)
            if ri==HDR_ROW:
                t.value=bcol;t.font=Font(name="Arial",size=10,bold=True,color="FFFFFF");t.fill=PatternFill("solid",fgColor=NAVY);t.alignment=Alignment(horizontal="center",vertical="center");t.border=BORDER
            elif ri==4 and cells and cells[0].value: cpc(cells[0],t)
            off=1
        for ci,src in enumerate(cells,start=1): cpc(src,ws.cell(ri,ci+off))
    r=HDR_ROW+1
    for cells,bnr in data:
        off=0
        if ins:
            t=ws.cell(r,1);t.value=bnr;ref=cells[0]
            t.font=Font(name="Arial",size=10,bold=True,color=ref.font.color.rgb if ref.font and ref.font.color and ref.font.color.rgb else "000000")
            if ref.fill and ref.fill.fgColor and ref.fill.fgColor.rgb: t.fill=copy(ref.fill)
            t.border=BORDER;off=1
        for ci,src in enumerate(cells,start=1): cpc(src,ws.cell(r,ci+off))
        r+=1
    last=r-1; r+=1
    for cells in tail:
        off=1 if ins else 0
        for ci,src in enumerate(cells,start=1): cpc(src,ws.cell(r,ci+off))
        r+=1
    ws.auto_filter.ref=f"A{HDR_ROW}:{get_column_letter(width)}{last}"; ws.freeze_panes=f"A{HDR_ROW+1}"
    for mr in list(ws.merged_cells.ranges): ws.unmerge_cells(str(mr))
    if ins: ws.column_dimensions["A"].width=22
    return last-HDR_ROW,len(tail)
wb=load_workbook(SRC)
for name,bcol in {"Portfolio + Factors":"Class","TradingView Watchlist":None,"Equity Master":None,"Industry":None}.items():
    if name in wb.sheetnames:
        n,t=flatten(wb[name],bcol); print(f"{name:22} data {n} | {t} below | {wb[name].auto_filter.ref}")
wb.save(DST); print("saved:",DST)
