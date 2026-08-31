#!/usr/bin/env python3
import sys, subprocess, os, shutil
from copy import copy
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
RAW=sys.argv[1]; OUT=sys.argv[2]; LBL_PREV=sys.argv[3]; LBL_CUR=sys.argv[4]; TMP="_fin_tmp.xlsx"
subprocess.run([sys.executable,os.path.join(os.path.dirname(os.path.abspath(__file__)),"make_filterable.py"),RAW,TMP],check=True)
wb=load_workbook(TMP)
def snap(ws,r,NC): return [dict(v=ws.cell(r,c).value,font=copy(ws.cell(r,c).font),fill=copy(ws.cell(r,c).fill),border=copy(ws.cell(r,c).border),align=copy(ws.cell(r,c).alignment),nf=ws.cell(r,c).number_format) for c in range(1,NC+1)]
def put(ws,r,snaps,c0=1):
    for i,s in enumerate(snaps):
        cell=ws.cell(r,c0+i); cell.value=s["v"];cell.font=s["font"];cell.fill=s["fill"];cell.border=s["border"];cell.alignment=s["align"];cell.number_format=s["nf"]
pf=wb["Portfolio + Factors"]
for r in (1,2):
    if pf.cell(r,2).value and not pf.cell(r,1).value: pf.cell(r,1).value=pf.cell(r,2).value; pf.cell(r,1).font=copy(pf.cell(r,2).font); pf.cell(r,2).value=None
if pf.cell(4,1).value=="POSITION" and pf.cell(4,2).value=="POSITION": pf.cell(4,1).value=None
band={pf.cell(4,c).value:c for c in range(1,pf.max_column+1) if pf.cell(4,c).value}
if "FUNDAMENTALS" in band and "TECHNICALS" in band and band["FUNDAMENTALS"]<band["TECHNICALS"]:
    f0=band["FUNDAMENTALS"]; t0=band["TECHNICALS"]; end=pf.max_column
    fund=list(range(f0,t0)); tech=list(range(t0,end+1)); OLD=list(range(f0,end+1)); neworder=tech+fund
    for r in range(1,pf.max_row+1):
        s=snap(pf,r,pf.max_column); put(pf,r,[s[c-1] for c in neworder],c0=f0)
    widths=[pf.column_dimensions[get_column_letter(c)].width for c in OLD]; wnew=[widths[OLD.index(c)] for c in neworder]
    for i,dst in enumerate(range(f0,end+1)): pf.column_dimensions[get_column_letter(dst)].width=wnew[i]
for n in ["Portfolio + Factors","Equity Master"]:
    ws=wb[n]; hdr=[str(ws.cell(5,c).value or "").strip() for c in range(1,ws.max_column+1)]
    qi=[i+1 for i,h in enumerate(hdr) if h.endswith("Qty") and not h.startswith("Δ")]
    if len(qi)>=2: ws.cell(5,qi[0]).value=f"{LBL_PREV} Qty"; ws.cell(5,qi[1]).value=f"{LBL_CUR} Qty"
if "Industry" in wb.sheetnames:
    ws=wb["Industry"]; NC=ws.max_column
    order=[]; gmap={}; subs={}; total=None
    for r in range(6,ws.max_row+1):
        c1=str(ws.cell(r,1).value or ""); c2=str(ws.cell(r,2).value or "")
        if c2.startswith("Sub-total"): subs[c1]=snap(ws,r,NC)
        elif "TOTAL EQUITY" in c1.upper() or "TOTAL EQUITY" in c2.upper(): total=snap(ws,r,NC)
        elif c1 or c2:
            if c1 not in gmap: gmap[c1]=[]; order.append(c1)
            gmap[c1].append(snap(ws,r,NC))
    ws.delete_rows(6, ws.max_row-5); r=6
    for ind in order:
        for row in gmap[ind]: put(ws,r,row); r+=1
        if ind in subs: put(ws,r,subs[ind]); r+=1
    last=r-1; r+=1
    if total: put(ws,r,total)
    ws.auto_filter.ref=f"A5:{get_column_letter(NC)}{last}"; ws.freeze_panes="A6"
wb.save(TMP)
d="_lo"; shutil.rmtree(d,ignore_errors=True); os.makedirs(d+"/out",exist_ok=True); shutil.copy(TMP,d+"/src.xlsx")
subprocess.run(["libreoffice","--headless","--calc","--convert-to","xlsx:Calc MS Excel 2007 XML","src.xlsx","--outdir","out"],cwd=d,capture_output=True,timeout=120)
shutil.copy(d+"/out/src.xlsx",OUT); os.remove(TMP); print("finalized ->",OUT)
