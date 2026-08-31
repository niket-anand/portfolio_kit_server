#!/usr/bin/env python3
"""Append P&L x Weekly-Trend grid + flat classification sheets to a review workbook.
Usage: python tools/pnl_grid.py "outputs/Weekly Portfolio Review <date>.xlsx" """
import sys, pickle
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
P=sys.argv[1]
NAVY="1F3864"; thin=Side(style="thin",color="D9D9D9"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
wb=load_workbook(P); em=wb["Equity Master"]; h={em.cell(5,c).value:c for c in range(1,em.max_column+1)}
rows=[]
for r in range(6,em.max_row+1):
    nm=em.cell(r,1).value
    if not nm: continue
    rows.append((nm, em.cell(r,h["NSE"]).value,
        pd.to_numeric(em.cell(r,h["P&L %"]).value,errors="coerce"),
        pd.to_numeric(em.cell(r,h["Wk Ret %"]).value,errors="coerce"),
        pd.to_numeric(em.cell(r,h["% of Port"]).value,errors="coerce")))
def pb(x):
    if pd.isna(x): return None
    return "P&L >20%" if x>=20 else "P&L 10-19%" if x>=10 else "P&L 0-10%" if x>=0 else "P&L -ve"
def tb(x):
    if pd.isna(x): return None
    return "Wk >5%" if x>5 else "Wk 0-5%" if x>=0 else "Wk 0 to -3%" if x>=-3 else "Wk <-3%"
Prows=["P&L >20%","P&L 10-19%","P&L 0-10%","P&L -ve"]; Tcols=["Wk >5%","Wk 0-5%","Wk 0 to -3%","Wk <-3%"]
names={(p,t):[] for p in Prows for t in Tcols}; recs=[]
for nm,nse,pnl,wk,wt in rows:
    p=pb(pnl); t=tb(wk); recs.append(dict(Name=nm,NSE=nse,PnL=pnl,Wk=wk,Wt=wt,P=p,T=t))
    if p and t: names[(p,t)].append(nm)
d=pd.DataFrame(recs)
def tint(pi,ti):
    palette=["F8696B","FCA36B","FFD666","D9E86B","A9D08E","7EC17E","63BE7B"]; return palette[(3-pi)+(3-ti)]
for s in ["P&L × Trend Grid","Grid — Classification"]:
    if s in wb.sheetnames: del wb[s]
ws=wb.create_sheet("P&L × Trend Grid"); ws.sheet_view.showGridLines=False
ws["A1"]="PORTFOLIO GRID — P&L × Weekly Trend"; ws["A1"].font=Font(name="Arial",size=13,bold=True,color=NAVY)
ws["A2"]=f"{sum(len(v) for v in names.values())} held names · rows=P&L%, cols=weekly return% · green=winning+rising"
ws["A2"].font=Font(name="Arial",size=9,italic=True,color="808080")
ws.cell(4,1,"P&L ⧵ Weekly").font=Font(name="Arial",size=10,bold=True,color="FFFFFF")
ws.cell(4,1).fill=PatternFill("solid",fgColor=NAVY); ws.cell(4,1).border=bd; ws.cell(4,1).alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
for j,t in enumerate(Tcols):
    x=ws.cell(4,2+j,t); x.font=Font(name="Arial",size=10,bold=True,color="FFFFFF"); x.fill=PatternFill("solid",fgColor=NAVY); x.border=bd; x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
ws.cell(4,6,"Row total").font=Font(name="Arial",size=10,bold=True,color="FFFFFF"); ws.cell(4,6).fill=PatternFill("solid",fgColor="808080"); ws.cell(4,6).border=bd; ws.cell(4,6).alignment=Alignment(horizontal="center")
r=5
for pi,p in enumerate(Prows):
    rl=ws.cell(r,1,p); rl.font=Font(name="Arial",size=10,bold=True,color="FFFFFF"); rl.fill=PatternFill("solid",fgColor=NAVY); rl.border=bd; rl.alignment=Alignment(horizontal="center",vertical="center")
    rt=0
    for ti,t in enumerate(Tcols):
        lst=names[(p,t)]; rt+=len(lst)
        cell=ws.cell(r,2+ti,(f"({len(lst)})  "+", ".join(lst)) if lst else "—"); cell.border=bd
        cell.font=Font(name="Arial",size=9); cell.fill=PatternFill("solid",fgColor=tint(pi,ti)); cell.alignment=Alignment(horizontal="left",vertical="top",wrap_text=True)
    tc=ws.cell(r,6,rt); tc.font=Font(name="Arial",size=10,bold=True); tc.border=bd; tc.alignment=Alignment(horizontal="center"); ws.row_dimensions[r].height=70; r+=1
ws.cell(r,1,"Col total").font=Font(name="Arial",size=10,bold=True,color="FFFFFF"); ws.cell(r,1).fill=PatternFill("solid",fgColor="808080"); ws.cell(r,1).border=bd; ws.cell(r,1).alignment=Alignment(horizontal="center")
for ti,t in enumerate(Tcols):
    x=ws.cell(r,2+ti,sum(len(names[(p,t)]) for p in Prows)); x.font=Font(name="Arial",size=10,bold=True); x.border=bd; x.alignment=Alignment(horizontal="center")
ws.cell(r,6,sum(len(v) for v in names.values())).font=Font(name="Arial",size=10,bold=True); ws.cell(r,6).border=bd; ws.cell(r,6).alignment=Alignment(horizontal="center")
ws.column_dimensions["A"].width=14
for col in "BCDE": ws.column_dimensions[col].width=34
ws.column_dimensions["F"].width=9
fs=wb.create_sheet("Grid — Classification"); fs.sheet_view.showGridLines=False
fs["A1"]="STOCK CLASSIFICATION — P&L & Weekly Trend"; fs["A1"].font=Font(name="Arial",size=12,bold=True,color=NAVY)
for c,hh in enumerate(["Company","NSE","P&L %","Wk Ret %","P&L Bucket","Trend Bucket","% of Port"],1):
    x=fs.cell(3,c,hh); x.font=Font(name="Arial",size=10,bold=True,color="FFFFFF"); x.fill=PatternFill("solid",fgColor=NAVY); x.border=bd; x.alignment=Alignment(horizontal="center",wrap_text=True)
rr=4
for _,x in d.sort_values(["P","Wk"],ascending=[True,False]).iterrows():
    for c,v in enumerate([x["Name"],x["NSE"],x["PnL"],x["Wk"],x["P"] or "—",x["T"] or "—",x["Wt"]],1):
        cell=fs.cell(rr,c,v); cell.border=bd; cell.font=Font(name="Arial",size=10)
        if c in (3,4,7) and pd.notna(v):
            cell.number_format='0.0"%"'
            if c in (3,4): cell.font=Font(name="Arial",size=10,color="9C0006" if v<0 else "006100")
    rr+=1
for col,w in zip("ABCDEFG",[22,12,9,10,14,14,10]): fs.column_dimensions[col].width=w
fs.freeze_panes="A4"; fs.auto_filter.ref=f"A3:G{rr-1}"
wb.save(P); print("grid sheets added ->",P)
