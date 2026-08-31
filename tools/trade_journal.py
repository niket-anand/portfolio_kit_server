#!/usr/bin/env python3
"""Build a Trade Journal xlsx from broker trade files in inputs/trades/.
Handles ICICI order books (orderBook_Equity_*.csv, tab of trades w/ 'DP Id - Client DP Id')
and Zerodha tradebooks (tradebook-*.csv, cols symbol/isin/trade_date/trade_type/quantity/price).
Optional month filter: python tools/trade_journal.py 2026-08
Zerodha partial fills are aggregated per symbol/day/side. ICICI DP->person mapping is
tentative (first DP seen -> Account 1, second -> Account 2) -- edit HOLDER_MAP below."""
import sys, glob, os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
MONTH=sys.argv[1] if len(sys.argv)>1 else None   # e.g. "2026-08" to filter
TR="inputs/trades"
HOLDER_MAP={}  # optionally map DP-Id -> person, e.g. {"IN301234-56780000":"Account 1"}
def rd(p):
    for e in ["utf-8","cp1252","latin-1"]:
        try: return pd.read_csv(p,encoding=e)
        except: pass
recs=[]; icici_dps=[]
for p in sorted(glob.glob(TR+"/*.csv")):
    d=rd(p);
    if d is None: continue
    d.columns=[c.strip() for c in d.columns]
    if "trade_type" in d.columns and "symbol" in d.columns:      # Zerodha
        d["val"]=pd.to_numeric(d["quantity"],errors="coerce")*pd.to_numeric(d["price"],errors="coerce")
        g=d.groupby(["symbol","isin","trade_date","trade_type"],as_index=False).agg(Qty=("quantity","sum"),Val=("val","sum"),Time=("order_execution_time","min"),fills=("price","size"))
        for _,r in g.iterrows():
            recs.append(dict(Date=str(r["trade_date"]),Holder="Zerodha Account",Broker="Zerodha",Symbol=str(r["symbol"]).strip(),
                ISIN=str(r["isin"]).strip(),Action=str(r["trade_type"]).strip().capitalize(),Qty=float(r["Qty"]),
                Price=round(r["Val"]/r["Qty"],2),TradeValue=round(r["Val"],2),Charges=None,Fills=int(r["fills"])))
    elif "DP Id - Client DP Id" in d.columns:                    # ICICI order book
        dp=str(d["DP Id - Client DP Id"].iloc[0]).strip(); icici_dps.append(dp)
        for _,r in d.iterrows():
            recs.append(dict(Date=pd.to_datetime(r["Date"],format="%d-%b-%y",errors="coerce").strftime("%Y-%m-%d"),
                Holder=None,Broker=f"ICICI ({dp})",Symbol=str(r["Stock"]).strip(),ISIN="",Action=str(r["Action"]).strip(),
                Qty=float(r["Qty"]),Price=round(float(r["Price"]),2),TradeValue=round(float(r["Trade Value"]),2),
                Charges=round(float(r.get("Brokerage Incl. Taxes",0) or 0),2),Fills=1,DP=dp))
# assign ICICI holders (tentative)
dps=sorted(set(icici_dps)); auto={dps[i]:f"Account {i+1}" for i in range(len(dps))}
for r in recs:
    if r["Holder"] is None:
        r["Holder"]=HOLDER_MAP.get(r.get("DP"), auto.get(r.get("DP"),"?"))
df=pd.DataFrame(recs)
if MONTH: df=df[df["Date"].astype(str).str.startswith(MONTH)]
df=df.sort_values(["Date","Holder","Symbol"]).reset_index(drop=True)
NAVY="1F3864"; thin=Side(style="thin",color="D9D9D9"); bd=Border(left=thin,right=thin,top=thin,bottom=thin)
GREEN="006100";GF="C6EFCE";RED="9C0006";RF="FFC7CE"
HDR=["Date","Holder","Broker / Account","Symbol","ISIN","Action","Qty","Avg Price (₹)","Trade Value (₹)","Charges (₹)","Net Cashflow (₹)","Fills"]
def net(r): ch=r["Charges"] or 0; return round(-(r["TradeValue"]+ch),2) if r["Action"]=="Buy" else round(r["TradeValue"]-ch,2)
def sheet(ws,data,title):
    ws.sheet_view.showGridLines=False; ws["A1"]=title; ws["A1"].font=Font(name="Arial",size=13,bold=True,color=NAVY)
    ws["A2"]=f"{len(data)} lines · Zerodha aggregates same-day fills · Buy=cash out, Sell=cash in · ICICI DP→person tentative"
    ws["A2"].font=Font(name="Arial",size=9,italic=True,color="808080")
    for c,hh in enumerate(HDR,1):
        x=ws.cell(4,c,hh); x.font=Font(name="Arial",size=10,bold=True,color="FFFFFF"); x.fill=PatternFill("solid",fgColor=NAVY); x.border=bd; x.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
    r=5
    for _,row in data.iterrows():
        for c,v in enumerate([row["Date"],row["Holder"],row["Broker"],row["Symbol"],row["ISIN"],row["Action"],row["Qty"],row["Price"],row["TradeValue"],row["Charges"],net(row),row["Fills"]],1):
            cell=ws.cell(r,c,v); cell.border=bd; cell.font=Font(name="Arial",size=10)
            if c==6: buy=row["Action"]=="Buy"; cell.font=Font(name="Arial",size=10,bold=True,color=GREEN if buy else RED); cell.fill=PatternFill("solid",fgColor=GF if buy else RF); cell.alignment=Alignment(horizontal="center")
            if c==7: cell.number_format='#,##0'
            if c in (8,9,10): cell.number_format='#,##0.00'
            if c==11: cell.number_format='#,##0;[Red]-#,##0'; cell.font=Font(name="Arial",size=10,color=RED if (v or 0)<0 else GREEN)
        r+=1
    for col,w in zip("ABCDEFGHIJKL",[11,11,22,12,15,7,8,13,15,11,15,6]): ws.column_dimensions[col].width=w
    ws.freeze_panes="A5"; ws.auto_filter.ref=f"A4:L{r-1}"
wb=Workbook(); sheet(wb.active,df,"TRADE JOURNAL"+(f" — {MONTH}" if MONTH else "")); wb.active.title="All Trades"
for hold in sorted(set(df["Holder"])-{None,"?"}):
    sheet(wb.create_sheet(hold),df[df["Holder"]==hold],f"TRADE JOURNAL — {hold}")
os.makedirs("outputs",exist_ok=True); wb.save("outputs/Trade Journal.xlsx")
print(f"Trade Journal: {len(df)} lines -> outputs/Trade Journal.xlsx")
